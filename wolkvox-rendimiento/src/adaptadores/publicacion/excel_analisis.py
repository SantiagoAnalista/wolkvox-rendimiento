"""Excel del informe de análisis: una hoja por tema, con índice, alertas
resaltadas y gráficos. Pensado para que un jefe de operación lo abra y
entienda sin que nadie se lo explique.
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

FUENTE = "Calibri"
AZUL = "1F3864"
AZUL_CLARO = "D9E2F3"
ROJO_FONDO, ROJO_TEXTO = "FFC7CE", "9C0006"
VERDE_FONDO, VERDE_TEXTO = "C6EFCE", "006100"
AMBAR_FONDO, AMBAR_TEXTO = "FFEB9C", "9C6500"

ENCABEZADO = Font(name=FUENTE, bold=True, color="FFFFFF", size=11)
FILL_ENCABEZADO = PatternFill("solid", fgColor=AZUL)
TITULO = Font(name=FUENTE, bold=True, size=14, color=AZUL)
BORDE = Border(*[Side(style="thin", color="BFBFBF")] * 4)

HOJAS = [
    ("Resumen general", "Promedios de toda la operación: puntualidad, tiempos y efectividad."),
    ("Resumen por agente", "Una fila por asesor con lo esencial de los tres frentes y sus alertas."),
    ("Puntualidad agente", "Entradas tarde y salidas temprano por asesor, con porcentajes."),
    ("Puntualidad detalle", "Día por día de cada asesor. La evidencia detrás de los porcentajes."),
    ("Tiempos por agente", "Totales por estado y mes, con % sobre el tiempo logueado y promedio diario."),
    ("Auxiliares detalle", "Tabla cruzada: horas y minutos de cada asesor en cada estado auxiliar."),
    ("Auxiliares día a día", "El mismo desglose fecha por fecha, para ver en qué días concretos se dispara."),
    ("Auxiliares por tipo", "Lo mismo en formato de lista, ordenado por duración dentro de cada asesor."),
    ("Efectividad", "Gestión efectiva por canal: voz y digital (WhatsApp/chat)."),
    ("Gestión vs efectividad", "Cruce de tiempo conectado contra resultados. Detecta 'conectado pero poco efectivo'."),
]


# Prefijo de todo informe publicado. Una sola definición: la usan la
# retención local, la publicación en red y sus expresiones regulares.
PREFIJO = "gestion_wolkbox"

def _escribir_tabla(ws, df: pd.DataFrame, fila_inicio: int = 1, filtro: bool = True) -> int:
    """Escribe el DataFrame con encabezado y devuelve la última fila usada."""
    if df.empty:
        ws.cell(row=fila_inicio, column=1, value="Sin datos para el periodo consultado").font = Font(
            name=FUENTE, italic=True, color="808080")
        return fila_inicio

    for r, fila in enumerate(dataframe_to_rows(df, index=False, header=True), start=fila_inicio):
        for c, valor in enumerate(fila, start=1):
            celda = ws.cell(row=r, column=c, value=valor)
            celda.border = BORDE
            if r == fila_inicio:
                celda.font = ENCABEZADO
                celda.fill = FILL_ENCABEZADO
                celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                celda.font = Font(name=FUENTE, size=10)
                if isinstance(valor, (int, float)) and not isinstance(valor, bool):
                    celda.alignment = Alignment(horizontal="center")

    fin = fila_inicio + len(df)
    for i, col in enumerate(df.columns, start=1):
        ancho = max(len(str(col)), *(len(str(v)) for v in df[col].head(200))) if len(df) else len(str(col))
        ws.column_dimensions[get_column_letter(i)].width = min(max(ancho + 3, 10), 34)
    ws.freeze_panes = ws.cell(row=fila_inicio + 1, column=1)
    if filtro:
        ws.auto_filter.ref = f"A{fila_inicio}:{get_column_letter(len(df.columns))}{fin}"
    return fin


def _col(df: pd.DataFrame, nombre: str) -> str | None:
    return get_column_letter(list(df.columns).index(nombre) + 1) if nombre in df.columns else None


def _semaforo(ws, df, columna, fila_ini, fila_fin, umbral_malo, invertir=False):
    """Pinta de rojo lo que supera el umbral (o lo que queda por debajo)."""
    letra = _col(df, columna)
    if not letra or fila_fin <= fila_ini:
        return
    rango = f"{letra}{fila_ini + 1}:{letra}{fila_fin}"
    operador, formula = ("lessThan", [str(umbral_malo)]) if invertir else ("greaterThan", [str(umbral_malo)])
    ws.conditional_formatting.add(rango, CellIsRule(
        operator=operador, formula=formula,
        fill=PatternFill("solid", fgColor=ROJO_FONDO), font=Font(color=ROJO_TEXTO, bold=True)))


def _escala(ws, df, columna, fila_ini, fila_fin, invertir=False):
    letra = _col(df, columna)
    if not letra or fila_fin <= fila_ini:
        return
    inicio, fin = (VERDE_FONDO, ROJO_FONDO) if invertir else (ROJO_FONDO, VERDE_FONDO)
    ws.conditional_formatting.add(
        f"{letra}{fila_ini + 1}:{letra}{fila_fin}",
        ColorScaleRule(start_type="min", start_color=inicio, end_type="max", end_color=fin))


def _grafico_barras(ws, df, col_valor: str, titulo: str, ancla: str, fila_fin: int,
                    fila_encabezado: int = 1):
    letra = _col(df, col_valor)
    if not letra or df.empty:
        return
    idx = list(df.columns).index(col_valor) + 1
    gr = BarChart()
    gr.title = titulo
    gr.add_data(Reference(ws, min_col=idx, min_row=fila_encabezado, max_row=fila_fin),
                titles_from_data=True)
    gr.set_categories(Reference(ws, min_col=1, min_row=fila_encabezado + 1, max_row=fila_fin))
    gr.height, gr.width, gr.legend = 9, 20, None
    ws.add_chart(gr, ancla)


def _hoja_indice(wb, metadatos: dict):
    ws = wb.create_sheet("Índice")
    ws.cell(row=1, column=1, value="Informe de gestión de asesores").font = Font(
        name=FUENTE, bold=True, size=18, color=AZUL)
    ws.cell(row=2, column=1, value=metadatos.get("Periodo", "")).font = Font(
        name=FUENTE, size=11, color="595959")

    ws.cell(row=4, column=1, value="Contenido").font = TITULO
    fila = 5
    for nombre, descripcion in HOJAS:
        c = ws.cell(row=fila, column=1, value=nombre)
        c.font = Font(name=FUENTE, bold=True, size=11, color="0563C1", underline="single")
        c.hyperlink = f"#'{nombre}'!A1"
        ws.cell(row=fila, column=2, value=descripcion).font = Font(name=FUENTE, size=10)
        fila += 1

    fila += 1
    ws.cell(row=fila, column=1, value="Criterios aplicados").font = TITULO
    fila += 1
    for clave, valor in metadatos.items():
        ws.cell(row=fila, column=1, value=clave).font = Font(name=FUENTE, bold=True, size=10)
        ws.cell(row=fila, column=2, value=str(valor)).font = Font(name=FUENTE, size=10)
        fila += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 95
    ws.sheet_view.showGridLines = False


def generar(cuadros: dict[str, pd.DataFrame], metadatos: dict, ruta_salida: Path,
            nombre: str = PREFIJO) -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    _hoja_indice(wb, metadatos)

    # --- Resumen general ---
    ws = wb.create_sheet("Resumen general")
    ws.cell(row=1, column=1, value="Puntualidad").font = TITULO
    fin_p = _escribir_tabla(ws, cuadros["general_puntualidad"], fila_inicio=2, filtro=False)
    ws.cell(row=fin_p + 2, column=1, value="Tiempos y efectividad").font = TITULO
    _escribir_tabla(ws, cuadros["general_gestion"], fila_inicio=fin_p + 3, filtro=False)
    ws.sheet_view.showGridLines = False

    # --- Resumen por agente ---
    df = cuadros["resumen_agente"]
    ws = wb.create_sheet("Resumen por agente")
    fin = _escribir_tabla(ws, df)
    _semaforo(ws, df, "% Entradas tarde", 1, fin, 10)
    _semaforo(ws, df, "% Salidas temprano", 1, fin, 10)
    _semaforo(ws, df, "% Auxiliar", 1, fin, 25)
    _semaforo(ws, df, "% Sin tipificar", 1, fin, 30)
    _escala(ws, df, "% Efectividad total", 1, fin, invertir=True)
    letra = _col(df, "Alerta")
    if letra and fin > 1:
        ws.conditional_formatting.add(
            f"A2:{get_column_letter(len(df.columns))}{fin}",
            CellIsRule(operator="containsText",
                       formula=[f'NOT(ISERROR(SEARCH("(",${letra}2)))'],
                       fill=PatternFill("solid", fgColor=AMBAR_FONDO),
                       font=Font(color=AMBAR_TEXTO)))

    # --- Puntualidad por agente ---
    df = cuadros["puntualidad_agente"]
    ws = wb.create_sheet("Puntualidad agente")
    fin = _escribir_tabla(ws, df)
    _semaforo(ws, df, "% Entradas tarde", 1, fin, 10)
    _semaforo(ws, df, "% Salidas temprano", 1, fin, 10)
    _semaforo(ws, df, "% Sin conexión", 1, fin, 10)
    _grafico_barras(ws, df, "% Entradas tarde", "% de entradas tarde por asesor",
                    f"A{fin + 3}", fin)

    # --- Puntualidad detalle ---
    df = cuadros["puntualidad_detalle"]
    ws = wb.create_sheet("Puntualidad detalle")
    fin = _escribir_tabla(ws, df)
    letra_estado = _col(df, "Estado")
    if letra_estado and fin > 1:
        rango = f"A2:{get_column_letter(len(df.columns))}{fin}"
        for texto, fondo, color in (("Tarde", ROJO_FONDO, ROJO_TEXTO),
                                    ("Temprano", AMBAR_FONDO, AMBAR_TEXTO),
                                    ("Sin conexión", AZUL_CLARO, AZUL)):
            ws.conditional_formatting.add(rango, CellIsRule(
                operator="containsText", formula=[f'NOT(ISERROR(SEARCH("{texto}",${letra_estado}2)))'],
                fill=PatternFill("solid", fgColor=fondo), font=Font(color=color)))

    # --- Tiempos por agente ---
    df = cuadros["tiempos"]
    ws = wb.create_sheet("Tiempos por agente")
    fin = _escribir_tabla(ws, df)
    _semaforo(ws, df, "% Auxiliar", 1, fin, 25)
    _escala(ws, df, "Ocupación %", 1, fin)
    _grafico_barras(ws, df, "% Auxiliar", "% de tiempo en auxiliares", f"A{fin + 3}", fin)

    # --- Auxiliares detalle (tabla cruzada) ---
    ws = wb.create_sheet("Auxiliares detalle")
    ws.cell(row=1, column=1, value="Horas y minutos por estado auxiliar").font = TITULO
    df_hm = cuadros["auxiliares_hm"]
    fin_hm = _escribir_tabla(ws, df_hm, fila_inicio=2, filtro=False)

    df_h = cuadros["auxiliares_horas"]
    ws.cell(row=fin_hm + 2, column=1,
            value="Lo mismo en horas decimales (para sumar y graficar)").font = TITULO
    fin_h = _escribir_tabla(ws, df_h, fila_inicio=fin_hm + 3, filtro=False)
    for columna in df_h.columns[1:]:
        _escala(ws, df_h, columna, fin_hm + 3, fin_h)

    if not df_h.empty and len(df_h.columns) > 1:
        _grafico_barras(ws, df_h, "TOTAL auxiliar", "Horas totales en auxiliares",
                        f"A{fin_h + 3}", fin_h, fila_encabezado=fin_hm + 3)
    ws.sheet_view.showGridLines = False

    # --- Auxiliares día a día ---
    df = cuadros["auxiliares_dia"]
    ws = wb.create_sheet("Auxiliares día a día")
    fin = _escribir_tabla(ws, df)

    # --- Auxiliares por tipo (lista) ---
    df = cuadros["auxiliares"]
    ws = wb.create_sheet("Auxiliares por tipo")
    fin = _escribir_tabla(ws, df)
    _escala(ws, df, "Horas", 1, fin)

    # --- Efectividad ---
    df = cuadros["efectividad"]
    ws = wb.create_sheet("Efectividad")
    fin = _escribir_tabla(ws, df)
    _escala(ws, df, "% Efectividad total", 1, fin, invertir=True)
    _grafico_barras(ws, df, "% Efectividad total", "% de efectividad por asesor",
                    f"A{fin + 3}", fin)

    # --- Cruce ---
    df = cuadros["cruce"]
    ws = wb.create_sheet("Gestión vs efectividad")
    fin = _escribir_tabla(ws, df)
    _semaforo(ws, df, "% Efectividad total", 1, fin, 0, invertir=False)
    _escala(ws, df, "% Efectividad total", 1, fin, invertir=True)
    _semaforo(ws, df, "% Auxiliar", 1, fin, 25)
    letra_alerta = _col(df, "Alerta")
    if letra_alerta and fin > 1:
        ws.conditional_formatting.add(
            f"A2:{get_column_letter(len(df.columns))}{fin}",
            CellIsRule(operator="containsText",
                       formula=[f'NOT(ISERROR(SEARCH("Conectado",${letra_alerta}2)))'],
                       fill=PatternFill("solid", fgColor=ROJO_FONDO),
                       font=Font(color=ROJO_TEXTO, bold=True)))

    ruta_salida.mkdir(parents=True, exist_ok=True)
    destino = ruta_salida / f"{nombre}.xlsx"
    wb.save(destino)
    return destino
