"""Arma los cuadros del reporte a partir de los DataFrames y los escribe en
Excel con gráficos nativos.

Los tiempos se muestran como 'HH:MM:SS' porque el Excel lo lee una persona;
los segundos crudos quedan en el backup CSV para cuando haya que calcular.
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

FUENTE = "Arial"
ENCABEZADO_FILL = PatternFill("solid", fgColor="1F3864")
ENCABEZADO_FONT = Font(name=FUENTE, bold=True, color="FFFFFF", size=11)


def _hhmmss(segundos) -> str:
    s = int(segundos or 0)
    return f"{s // 3600:02d}:{s % 3600 // 60:02d}:{s % 60:02d}"


def _pct(serie: pd.Series) -> pd.Series:
    """La ocupación ya viene en porcentaje (0-100) desde Wolkvox."""
    return serie.astype(float).round(1)


def resumen(dfs: dict[str, pd.DataFrame]) -> pd.DataFrame:
    ag, llam, no_con = dfs["agente_dia"], dfs["llamadas"], dfs["no_conectadas"]
    filas = [
        ("Agentes con actividad", 0 if ag.empty else int(ag["agent_id"].nunique())),
        ("Llamadas gestionadas", 0 if ag.empty else int(ag["calls"].sum())),
        ("  Entrantes", 0 if ag.empty else int(ag["inbound"].sum())),
        ("  Salientes", 0 if ag.empty else int(ag["outbound"].sum())),
        ("  Internas", 0 if ag.empty else int(ag["internal"].sum())),
        ("Intentos no conectados", 0 if no_con.empty else len(no_con)),
        ("Llamadas con detalle (CDR)", 0 if llam.empty else len(llam)),
        ("AHT promedio", _hhmmss(0 if ag.empty else ag["aht_seg"].mean())),
        ("Ocupación promedio", f"{0.0 if ag.empty else round(ag['ocupacion'].mean(), 1)} %"),
        ("Hits", 0 if ag.empty else int(ag["hits"].sum())),
        ("RPC", 0 if ag.empty else int(ag["rpc"].sum())),
    ]
    return pd.DataFrame(filas, columns=["Indicador", "Valor"])


def por_agente(dfs: dict[str, pd.DataFrame]) -> pd.DataFrame:
    df = dfs["agente_dia"]
    if df.empty:
        return pd.DataFrame()
    salida = pd.DataFrame({
        "Agente": df["agent_name"],
        "Llamadas": df["calls"],
        "Entrantes": df["inbound"],
        "Salientes": df["outbound"],
        "Internas": df["internal"],
        "AHT": df["aht_seg"].map(_hhmmss),
        "ACW": df["acw_time_seg"].map(_hhmmss),
        "Logueado": df["login_time_seg"].map(_hhmmss),
        "Auxiliar": df["aux_time_seg"].map(_hhmmss),
        "Ocupación %": _pct(df["ocupacion"]),
        "Hits": df["hits"],
        "RPC": df["rpc"],
    })
    return salida.sort_values("Llamadas", ascending=False).reset_index(drop=True)


def por_hora(dfs: dict[str, pd.DataFrame]) -> pd.DataFrame:
    df = dfs["agente_hora"]
    if df.empty:
        return pd.DataFrame()
    agrupado = df.groupby("hora", as_index=False).agg(
        Llamadas=("calls", "sum"),
        Ocupacion=("ocupacion", "mean"),
    )
    agrupado["Ocupación %"] = _pct(agrupado["Ocupacion"])
    return agrupado[["hora", "Llamadas", "Ocupación %"]].rename(columns={"hora": "Hora"})


def por_categoria(dfs: dict[str, pd.DataFrame]) -> pd.DataFrame:
    partes = [df["categoria_negocio"] for df in (dfs["llamadas"], dfs["no_conectadas"]) if not df.empty]
    if not partes:
        return pd.DataFrame()
    conteo = pd.concat(partes).value_counts().reset_index()
    conteo.columns = ["Categoría", "Cantidad"]
    return conteo


def _escribir(ws, df: pd.DataFrame) -> None:
    for r, fila in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c, valor in enumerate(fila, start=1):
            celda = ws.cell(row=r, column=c, value=valor)
            celda.font = Font(name=FUENTE, size=10)
            if r == 1:
                celda.font = ENCABEZADO_FONT
                celda.fill = ENCABEZADO_FILL
                celda.alignment = Alignment(horizontal="center", vertical="center")

    for col in ws.columns:
        ancho = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(ancho + 3, 40)
    ws.freeze_panes = "A2"


def _grafico(gr, ws, df: pd.DataFrame, titulo: str, col_valor: int, ancla: str) -> None:
    n = len(df)
    gr.title = titulo
    gr.add_data(Reference(ws, min_col=col_valor, min_row=1, max_row=n + 1), titles_from_data=True)
    gr.set_categories(Reference(ws, min_col=1, min_row=2, max_row=n + 1))
    gr.height, gr.width = 10, 22
    ws.add_chart(gr, ancla)


def generar(dfs: dict[str, pd.DataFrame], metadatos: dict, ruta_salida: Path) -> Path:
    cuadros = {
        "Resumen": resumen(dfs),
        "Resultados": por_categoria(dfs),
        "Por agente": por_agente(dfs),
        "Por hora": por_hora(dfs),
    }

    wb = Workbook()
    wb.remove(wb.active)

    for titulo, df in cuadros.items():
        ws = wb.create_sheet(titulo)
        if df.empty:
            ws.cell(row=1, column=1, value="Sin datos para la ventana consultada")
            continue
        _escribir(ws, df)

        if titulo == "Resultados":
            _grafico(PieChart(), ws, df, "Distribución por categoría", 2, "E2")
        elif titulo == "Por agente":
            _grafico(BarChart(), ws, df, "Llamadas por agente", 2, f"A{len(df) + 4}")
        elif titulo == "Por hora":
            _grafico(LineChart(), ws, df, "Curva horaria de llamadas", 2, f"A{len(df) + 4}")

    ws_m = wb.create_sheet("Metadatos")
    for i, (clave, valor) in enumerate(metadatos.items(), start=1):
        ws_m.cell(row=i, column=1, value=clave).font = Font(name=FUENTE, bold=True)
        ws_m.cell(row=i, column=2, value=str(valor)).font = Font(name=FUENTE)
    ws_m.column_dimensions["A"].width = 28
    ws_m.column_dimensions["B"].width = 60

    ruta_salida.mkdir(parents=True, exist_ok=True)
    destino = ruta_salida / f"rendimiento_{datetime.now():%Y%m%d_%H%M}.xlsx"
    wb.save(destino)
    wb.save(ruta_salida / "rendimiento_ultimo.xlsx")  # ruta fija para "el último"
    return destino
