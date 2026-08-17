import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from src.adaptadores.publicacion import excel_operativo as report
from src.adaptadores.wolkvox import traduccion as transform

CATEGORIAS = {"conectadas": {"VENTA": "efectiva"}, "no_conectadas": {"busy": "no_contactada"}}

CRUDOS = {
    "agentes": [{"agent_id": "1", "agent_name": "Ana"}],
    "agente_dia": [
        {"agent_id": "1", "agent_name": "Ana", "calls": "8", "inbound": "5", "outbound": "3",
         "login_time": "01:00:00", "ready_time": "00:12:00", "aux_time": "00:06:00",
         "acw_time": "00:05:00", "aht": "00:04:00", "occupancy": "70.00 %",
         "hits": "2", "rpc": "1"},
        {"agent_id": "2", "agent_name": "Luis", "calls": "3", "inbound": "3",
         "login_time": "01:00:00", "ready_time": "00:30:00", "aht": "00:06:00",
         "occupancy": "50.00 %"},
    ],
    "agente_hora": [
        {"agent_id": "1", "date": "2026-08-11", "hour": "8", "calls": "5",
         "login_time": "01:00:00", "occupancy": "60.00 %"},
        {"agent_id": "1", "date": "2026-08-11", "hour": "9", "calls": "3",
         "login_time": "01:00:00", "occupancy": "40.00 %"},
    ],
    "llamadas": [
        {"conn_id": "c1", "cod_act": "VENTA", "date": "2026-08-11 08:15:00", "time_seg": "120"},
        {"conn_id": "c2", "cod_act": "OTRO", "date": "2026-08-11 09:20:00", "time_seg": "60"},
    ],
    "no_conectadas": [{"conn_id": "n1", "result": "Busy", "date": "2026-08-11 09:30:00"}],
}


def _dfs(crudos=None):
    crudos = crudos or CRUDOS
    return {
        "agentes": transform.agentes(crudos["agentes"]),
        "agente_dia": transform.agente_dia(crudos["agente_dia"], "2026-08-11"),
        "agente_hora": transform.agente_hora(crudos["agente_hora"], "2026-08-11"),
        "llamadas": transform.llamadas(crudos["llamadas"], CATEGORIAS),
        "no_conectadas": transform.llamadas_no_conectadas(crudos["no_conectadas"], CATEGORIAS),
    }


def test_genera_excel_con_todas_las_hojas():
    salida = Path(tempfile.mkdtemp())
    destino = report.generar(_dfs(), {"Fecha procesada": "2026-08-11"}, salida)

    assert destino.exists()
    assert (salida / "rendimiento_ultimo.xlsx").exists()
    hojas = load_workbook(destino).sheetnames
    assert hojas == ["Resumen", "Resultados", "Por agente", "Por hora", "Metadatos"]


def test_resumen_suma_las_llamadas():
    resumen = report.resumen(_dfs()).set_index("Indicador")["Valor"]
    assert resumen["Llamadas gestionadas"] == 11  # 8 + 3
    assert resumen["Agentes con actividad"] == 2
    assert resumen["Intentos no conectados"] == 1


def test_por_agente_ordena_por_llamadas():
    df = report.por_agente(_dfs())
    assert list(df["Agente"]) == ["Ana", "Luis"]
    assert df.loc[0, "AHT"] == "00:04:00"


def test_por_hora_agrupa_por_hora():
    df = report.por_hora(_dfs())
    assert list(df["Hora"]) == [8, 9]
    assert list(df["Llamadas"]) == [5, 3]
    assert list(df["Ocupación %"]) == [60.0, 40.0]


def test_ocupacion_llega_al_excel_como_porcentaje_de_wolkvox():
    assert list(report.por_agente(_dfs())["Ocupación %"]) == [70.0, 50.0]


def test_categorias_mezcla_conectadas_y_no_conectadas():
    df = report.por_categoria(_dfs()).set_index("Categoría")["Cantidad"]
    assert df["efectiva"] == 1
    assert df["sin_clasificar"] == 1
    assert df["no_contactada"] == 1


def test_dia_sin_datos_genera_excel_sin_romper():
    """Fuera de jornada las fuentes vienen vacías: el libro se arma igual."""
    vacios = {k: [] for k in CRUDOS}
    salida = Path(tempfile.mkdtemp())
    destino = report.generar(_dfs(vacios), {"Fecha procesada": "2026-08-11"}, salida)

    hoja = load_workbook(destino)["Por agente"]
    assert hoja.cell(row=1, column=1).value == "Sin datos para la ventana consultada"
